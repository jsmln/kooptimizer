import io
from PIL import Image
import gzip
import mimetypes
from io import BytesIO
import subprocess
import tempfile
import os

# Max attachment size per send (4 MB - Brevo API limit for attachments)
MAX_ATTACHMENT_SIZE = 4 * 1024 * 1024

def compress_pdf(pdf_bytes):
    """
    Compress PDF file using PyPDF2/pypdf to reduce file size.
    Returns compressed PDF bytes or None if compression fails.
    """
    try:
        try:
            # Try PyPDF2 first (older, more common)
            from PyPDF2 import PdfReader, PdfWriter
            
            pdf_input = BytesIO(pdf_bytes)
            reader = PdfReader(pdf_input)
            writer = PdfWriter()
            
            # Copy all pages
            for page in reader.pages:
                # Compress page content
                page.compress_content_streams()
                writer.add_page(page)
            
            # Write compressed PDF
            output = BytesIO()
            writer.write(output)
            output.seek(0)
            compressed = output.getvalue()
            
            # Only return if actually smaller
            if len(compressed) < len(pdf_bytes):
                return compressed
            return pdf_bytes
            
        except ImportError:
            # Try pypdf (newer fork) via dynamic import to avoid static analysis/import resolution issues
            import importlib
            pypdf = importlib.import_module('pypdf')
            PdfReader = getattr(pypdf, 'PdfReader')
            PdfWriter = getattr(pypdf, 'PdfWriter')
            
            pdf_input = BytesIO(pdf_bytes)
            reader = PdfReader(pdf_input)
            writer = PdfWriter()
            
            for page in reader.pages:
                page.compress_content_streams()
                writer.add_page(page)
            
            output = BytesIO()
            writer.write(output)
            output.seek(0)
            compressed = output.getvalue()
            
            if len(compressed) < len(pdf_bytes):
                return compressed
            return pdf_bytes
            
    except Exception as e:
        print(f"PDF compression error: {e}")
        return None

def convert_docx_to_pdf(docx_bytes):
    """
    Convert DOCX bytes to PDF using multiple fallback methods.
    Priority: pypandoc > LibreOffice > ReportLab
    Returns PDF bytes or None if conversion fails.
    """
    try:
        # Method 1: Try pypandoc (if available)
        try:
            # Dynamically import pypandoc to avoid static analysis errors when it's not installed.
            import importlib
            pypandoc = importlib.import_module('pypandoc')
            
            with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as temp_input:
                temp_input.write(docx_bytes)
                temp_input_path = temp_input.name
            
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_output:
                temp_output_path = temp_output.name
            
            try:
                # Use pypandoc to convert
                pypandoc.convert_file(
                    temp_input_path, 
                    'pdf', 
                    outputfile=temp_output_path,
                    extra_args=['--pdf-engine=wkhtmltopdf']
                )
                
                if os.path.exists(temp_output_path):
                    with open(temp_output_path, 'rb') as f:
                        pdf_bytes = f.read()
                    return pdf_bytes
            finally:
                if os.path.exists(temp_input_path):
                    os.remove(temp_input_path)
                if os.path.exists(temp_output_path):
                    os.remove(temp_output_path)
        except (ImportError, OSError, RuntimeError):
            # pypandoc not available, continue to next method
            pass
        
        # Method 2: Try LibreOffice
        with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as temp_input:
            temp_input.write(docx_bytes)
            temp_input_path = temp_input.name
        
        try:
            # Try using soffice (LibreOffice) for conversion
            subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', os.path.dirname(temp_input_path),
                temp_input_path
            ], check=True, timeout=30, capture_output=True)
            
            # LibreOffice creates file with same name but .pdf extension
            libreoffice_output = temp_input_path.rsplit('.', 1)[0] + '.pdf'
            if os.path.exists(libreoffice_output):
                with open(libreoffice_output, 'rb') as f:
                    pdf_bytes = f.read()
                os.remove(libreoffice_output)
                return pdf_bytes
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError):
            pass
        finally:
            if os.path.exists(temp_input_path):
                os.remove(temp_input_path)
        
        # Method 3: Fallback to ReportLab (text extraction only)
        return convert_with_reportlab(docx_bytes, 'docx')
                
    except Exception as e:
        print(f"DOCX to PDF conversion failed: {e}")
        import traceback
        traceback.print_exc()
        return None


def convert_xlsx_to_pdf(xlsx_bytes):
    """
    Convert XLSX bytes to PDF using multiple fallback methods.
    Automatically selects optimal page size from Legal, Letter, A4, A3, A2 to fit all content.
    Priority: openpyxl+reportlab > LibreOffice > basic fallback
    Returns PDF bytes or None if conversion fails.
    """
    try:
        # Method 1: Try openpyxl with reportlab for better formatting
        try:
            from openpyxl import load_workbook
            from reportlab.lib.pagesizes import A2, A3, A4, LETTER, LEGAL, landscape, portrait
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            xlsx_file = BytesIO(xlsx_bytes)
            wb = load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            
            # Extract all data
            data = []
            max_col_widths = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                data.append(row_data)
                
                # Track max width for each column
                for idx, cell in enumerate(row_data):
                    if idx >= len(max_col_widths):
                        max_col_widths.append(0)
                    max_col_widths[idx] = max(max_col_widths[idx], len(cell))
            
            if not data:
                return None
            
            num_cols = len(data[0]) if data else 0
            num_rows = len(data)
            
            # Determine best page size and orientation based on content
            # Available page sizes in order of size (width in portrait mode):
            # LETTER: 8.5" x 11" (612 x 792 pts)
            # LEGAL: 8.5" x 14" (612 x 1008 pts)  
            # A4: 8.27" x 11.69" (595 x 842 pts)
            # A3: 11.69" x 16.54" (842 x 1191 pts)
            # A2: 16.54" x 23.39" (1191 x 1684 pts)
            
            # Decision matrix based on columns and rows
            if num_cols <= 4:
                # Very narrow - portrait is fine
                pagesize = portrait(LETTER)
            elif num_cols <= 6:
                # Narrow - portrait A4 or landscape Letter
                pagesize = landscape(LETTER) if num_rows > 30 else portrait(A4)
            elif num_cols <= 8:
                # Medium - landscape Letter/A4
                pagesize = landscape(LEGAL) if num_rows > 40 else landscape(LETTER)
            elif num_cols <= 12:
                # Wide - landscape Legal or A3
                pagesize = landscape(A4) if num_rows < 20 else landscape(LEGAL)
            elif num_cols <= 16:
                # Very wide - A3 landscape
                pagesize = landscape(A3)
            elif num_cols <= 24:
                # Extremely wide - A2 landscape
                pagesize = landscape(A2)
            else:
                # Ultra wide - largest available
                pagesize = landscape(A2)
            
            # Create PDF with minimal margins to maximize content area
            pdf_file = BytesIO()
            doc = SimpleDocTemplate(
                pdf_file, 
                pagesize=pagesize,
                leftMargin=0.3*inch,
                rightMargin=0.3*inch,
                topMargin=0.3*inch,
                bottomMargin=0.3*inch
            )
            
            # Calculate available width and distribute to columns
            page_width = pagesize[0] - 0.6*inch  # Total width minus margins
            
            # Distribute width based on content length (weighted by character count)
            total_weight = sum(max_col_widths)
            col_widths = []
            
            if total_weight > 0:
                for width in max_col_widths:
                    # Proportional width based on content, with minimum
                    proportional = (width / total_weight) * page_width
                    col_widths.append(max(proportional, 0.35*inch))  # Minimum 0.35 inch
            else:
                # Equal distribution if no weight data
                col_widths = [page_width / num_cols] * num_cols
            
            # Adjust if total exceeds page width
            total_width = sum(col_widths)
            if total_width > page_width:
                scale_factor = page_width / total_width
                col_widths = [w * scale_factor for w in col_widths]
            
            # Determine font size based on column count and page size
            if num_cols > 20:
                header_font_size = 6
                data_font_size = 5
            elif num_cols > 12:
                header_font_size = 7
                data_font_size = 6
            elif num_cols > 8:
                header_font_size = 8
                data_font_size = 7
            else:
                header_font_size = 9
                data_font_size = 8
            
            # Create table with calculated widths
            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                
                # Data rows styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                
                # Grid lines - all borders
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                
                # Alternating row colors for readability
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
                
                # Word wrap
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            
            doc.build([table])
            pdf_file.seek(0)
            return pdf_file.getvalue()
            
        except Exception as e:
            print(f"openpyxl conversion attempt failed: {e}")
            import traceback
            traceback.print_exc()
        
        # Method 2: Try LibreOffice
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_input:
            temp_input.write(xlsx_bytes)
            temp_input_path = temp_input.name
        
        try:
            # Try using soffice (LibreOffice) for conversion
            subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', os.path.dirname(temp_input_path),
                temp_input_path
            ], check=True, timeout=30, capture_output=True)
            
            libreoffice_output = temp_input_path.rsplit('.', 1)[0] + '.pdf'
            if os.path.exists(libreoffice_output):
                with open(libreoffice_output, 'rb') as f:
                    pdf_bytes = f.read()
                os.remove(libreoffice_output)
                return pdf_bytes
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError):
            pass
        finally:
            if os.path.exists(temp_input_path):
                os.remove(temp_input_path)
        
        # Method 3: Basic reportlab fallback
        return convert_with_reportlab(xlsx_bytes, 'xlsx')
                
    except Exception as e:
        print(f"XLSX to PDF conversion failed: {e}")
        import traceback
        traceback.print_exc()
        return None


def convert_with_reportlab(file_bytes, file_type):
    """
    Fallback conversion using reportlab (text-only, basic).
    Supports multiple page sizes: Letter, Legal, A4, A3, A2.
    """
    try:
        from reportlab.lib.pagesizes import A2, A3, A4, LETTER, LEGAL, landscape, portrait
        from reportlab.platypus import SimpleDocTemplate, Paragraph, PageBreak, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        
        pdf_file = BytesIO()
        
        if file_type == 'xlsx':
            from openpyxl import load_workbook
            xlsx_file = BytesIO(file_bytes)
            wb = load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            
            # Extract data
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else "" for cell in row])
            
            if not data:
                return None
            
            num_cols = len(data[0])
            num_rows = len(data)
            
            # Select page size based on content
            if num_cols <= 6:
                pagesize = portrait(LETTER)
            elif num_cols <= 10:
                pagesize = landscape(LEGAL)
            elif num_cols <= 16:
                pagesize = landscape(A3)
            else:
                pagesize = landscape(A2)
            
            doc = SimpleDocTemplate(
                pdf_file,
                pagesize=pagesize,
                leftMargin=0.3*inch,
                rightMargin=0.3*inch,
                topMargin=0.3*inch,
                bottomMargin=0.3*inch
            )
            
            # Calculate column widths
            page_width = pagesize[0] - 0.6*inch
            col_widths = [page_width / num_cols] * num_cols
            
            # Dynamic font sizing
            font_size = 5 if num_cols > 20 else (6 if num_cols > 12 else 7)
            
            # Create simple table
            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), font_size),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('WORDWRAP', (0, 0), (-1, -1), True),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
            ]))
            
            doc.build([table])
                            
        elif file_type == 'docx':
            from docx import Document
            docx_file = BytesIO(file_bytes)
            doc_docx = Document(docx_file)
            
            # Use Letter size for documents (standard in US and widely compatible)
            doc = SimpleDocTemplate(
                pdf_file,
                pagesize=LETTER,
                leftMargin=0.75*inch,
                rightMargin=0.75*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            styles = getSampleStyleSheet()
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                spaceBefore=6,
                spaceAfter=6,
            )
            
            story = []
            for para in doc_docx.paragraphs:
                if para.text.strip():
                    # Convert to reportlab paragraph to handle line wrapping
                    p = Paragraph(para.text, normal_style)
                    story.append(p)
            
            doc.build(story)
        
        pdf_file.seek(0)
        return pdf_file.getvalue()
    except Exception as e:
        print(f"ReportLab fallback conversion failed: {e}")
        import traceback
        traceback.print_exc()
        return None

def convert_pptx_to_pdf(pptx_bytes):
    """
    Convert PPTX bytes to PDF using LibreOffice or fallback.
    Returns PDF bytes or None if conversion fails.
    """
    try:
        # Method 1: Try LibreOffice (best quality for presentations)
        with tempfile.NamedTemporaryFile(suffix='.pptx', delete=False) as temp_input:
            temp_input.write(pptx_bytes)
            temp_input_path = temp_input.name
        
        try:
            subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', os.path.dirname(temp_input_path),
                temp_input_path
            ], check=True, timeout=30, capture_output=True)
            
            libreoffice_output = temp_input_path.rsplit('.', 1)[0] + '.pdf'
            if os.path.exists(libreoffice_output):
                with open(libreoffice_output, 'rb') as f:
                    pdf_bytes = f.read()
                os.remove(libreoffice_output)
                return pdf_bytes
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError):
            pass
        finally:
            if os.path.exists(temp_input_path):
                os.remove(temp_input_path)
        
        # Method 2: Try python-pptx with reportlab fallback
        try:
            # Dynamically import python-pptx to avoid static analysis/import resolution issues
            import importlib
            try:
                pptx_module = importlib.import_module('pptx')
                Presentation = getattr(pptx_module, 'Presentation')
            except Exception:
                # If python-pptx is not available, bail out of this fallback
                raise ImportError("python-pptx not available")
            
            from reportlab.lib.pagesizes import landscape, LETTER
            from reportlab.pdfgen import canvas
            
            pptx_file = BytesIO(pptx_bytes)
            prs = Presentation(pptx_file)
            
            pdf_file = BytesIO()
            c = canvas.Canvas(pdf_file, pagesize=landscape(LETTER))
            width, height = landscape(LETTER)
            
            for slide_num, slide in enumerate(prs.slides):
                # Extract text from slide
                texts = []
                for shape in slide.shapes:
                    if hasattr(shape, "text"):
                        texts.append(shape.text)
                
                # Draw slide number
                c.setFont("Helvetica-Bold", 16)
                c.drawString(40, height - 40, f"Slide {slide_num + 1}")
                
                # Draw text content
                c.setFont("Helvetica", 12)
                y = height - 80
                for text in texts:
                    if text.strip():
                        c.drawString(40, y, text[:100])  # Truncate long lines
                        y -= 20
                        if y < 40:
                            break
                
                c.showPage()
            
            c.save()
            pdf_file.seek(0)
            return pdf_file.getvalue()
            
        except Exception as e:
            print(f"python-pptx fallback failed: {e}")
            return None
                
    except Exception as e:
        print(f"PPTX to PDF conversion failed: {e}")
        import traceback
        traceback.print_exc()
        return None


def convert_txt_to_pdf(txt_bytes):
    """
    Convert plain text to PDF with proper formatting.
    Returns PDF bytes or None if conversion fails.
    """
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        
        # Decode text
        try:
            text = txt_bytes.decode('utf-8')
        except:
            text = txt_bytes.decode('latin-1', errors='ignore')
        
        pdf_file = BytesIO()
        doc = SimpleDocTemplate(
            pdf_file,
            pagesize=LETTER,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Split by paragraphs
        paragraphs = text.split('\n')
        for para_text in paragraphs:
            if para_text.strip():
                p = Paragraph(para_text, styles['Normal'])
                story.append(p)
            else:
                story.append(Spacer(1, 0.1*inch))
        
        doc.build(story)
        pdf_file.seek(0)
        return pdf_file.getvalue()
        
    except Exception as e:
        print(f"TXT to PDF conversion failed: {e}")
        return None


def can_convert_to_pdf(content_type):
    """Check if file type can be converted to PDF."""
    convertible_types = [
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',  # .docx
        'application/msword',  # .doc
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
        'application/vnd.openxmlformats-officedocument.presentationml.presentation',  # .pptx
        'application/vnd.ms-powerpoint',  # .ppt
        'text/plain',  # .txt
        'text/csv',  # .csv (will be treated as text)
    ]
    return content_type in convertible_types

def convert_to_pdf(file_bytes, filename, content_type):
    """
    Convert office documents and text files to PDF.
    Returns (pdf_bytes, success) tuple.
    """
    try:
        # Handle memoryview objects
        if isinstance(file_bytes, memoryview):
            file_bytes = bytes(file_bytes)
        
        if 'wordprocessingml' in content_type or content_type == 'application/msword':
            pdf_bytes = convert_docx_to_pdf(file_bytes)
            if pdf_bytes:
                return pdf_bytes, True
        elif 'spreadsheet' in content_type or content_type == 'application/vnd.ms-excel':
            pdf_bytes = convert_xlsx_to_pdf(file_bytes)
            if pdf_bytes:
                return pdf_bytes, True
        elif 'presentationml' in content_type or content_type == 'application/vnd.ms-powerpoint':
            pdf_bytes = convert_pptx_to_pdf(file_bytes)
            if pdf_bytes:
                return pdf_bytes, True
        elif content_type in ['text/plain', 'text/csv']:
            pdf_bytes = convert_txt_to_pdf(file_bytes)
            if pdf_bytes:
                return pdf_bytes, True
    except Exception as e:
        print(f"Error in convert_to_pdf: {e}")
        import traceback
        traceback.print_exc()
    
    return None, False

def process_attachment(file_obj, filename):
    """
    Accepts a Django UploadedFile (file_obj) and filename.
    Compresses ALL file types to reduce database storage:
    - Images: resize/re-encode with optimized compression
    - PDFs: apply PDF compression techniques
    - Office docs (Word, Excel, PowerPoint): compress as-is
    - All other files: gzip compression
    Returns tuple: (bytes_data, content_type, final_filename, size)
    """
    # Read raw bytes
    raw = file_obj.read()
    size = len(raw)

    # Quick reject if already over hard limit (16MB absolute maximum)
    if size > MAX_ATTACHMENT_SIZE * 4:
        raise ValueError('File too large (maximum 16MB)')

    # Get content type
    provided_ct = getattr(file_obj, 'content_type', None)
    if not provided_ct:
        provided_ct, _ = mimetypes.guess_type(filename)
    
    # Try to detect image using Pillow
    is_image = False
    try:
        img_test = Image.open(io.BytesIO(raw))
        img_test.verify()
        is_image = True
    except Exception:
        pass
    
    if is_image:
        # ============ IMAGE COMPRESSION ============
        img = Image.open(io.BytesIO(raw))
        img_format = img.format or 'JPEG'

        # Convert palette mode to RGB for better compatibility
        if img.mode == 'P':
            img = img.convert('RGB')

        # Resize if width > 1920 or height > 1920
        max_dim = 1920
        if max(img.size) > max_dim:
            ratio = max_dim / float(max(img.size))
            new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
            img = img.resize(new_size, Image.LANCZOS)

        out = io.BytesIO()
        
        # Handle different image formats with optimized settings
        if img_format == 'PNG' and img.mode in ('RGBA', 'LA'):
            # For PNG with transparency, keep as PNG but optimize
            img.save(out, format='PNG', optimize=True, compress_level=9)
            data = out.getvalue()
            content_type = 'image/png'
            final_filename = filename.rsplit('.', 1)[0] + '.png'
        else:
            # Convert to JPEG with progressive encoding and optimal quality
            if img.mode in ('RGBA', 'LA', 'P'):
                # Convert to RGB with white background
                bg = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'RGBA' or img.mode == 'LA':
                    bg.paste(img, mask=img.split()[-1])
                else:
                    bg.paste(img)
                img = bg

            # Use progressive JPEG for better compression and web performance
            img.save(out, format='JPEG', quality=85, optimize=True, progressive=True)
            data = out.getvalue()
            content_type = 'image/jpeg'
            final_filename = filename.rsplit('.', 1)[0] + '.jpg'

        final_size = len(data)
        
        # If still too large, reduce quality further
        if final_size > MAX_ATTACHMENT_SIZE:
            out = io.BytesIO()
            if content_type == 'image/png':
                # For PNG, try converting to JPEG instead
                if img.mode in ('RGBA', 'LA'):
                    bg = Image.new('RGB', img.size, (255, 255, 255))
                    bg.paste(img, mask=img.split()[-1])
                    img = bg
                img.save(out, format='JPEG', quality=75, optimize=True, progressive=True)
                content_type = 'image/jpeg'
                final_filename = filename.rsplit('.', 1)[0] + '.jpg'
            else:
                img.save(out, format='JPEG', quality=70, optimize=True, progressive=True)
            data = out.getvalue()
            final_size = len(data)

        if final_size > MAX_ATTACHMENT_SIZE:
            raise ValueError('Image could not be reduced below 4MB')

        return data, content_type, final_filename, final_size

    else:
        # ============ NON-IMAGE FILE COMPRESSION ============
        
        # Check if it's a PDF
        is_pdf = provided_ct == 'application/pdf' or filename.lower().endswith('.pdf')
        
        # Check if it's an Office document (already compressed internally)
        office_types = [
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',  # .docx
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.openxmlformats-officedocument.presentationml.presentation',  # .pptx
            'application/msword',  # .doc
            'application/vnd.ms-excel',  # .xls
            'application/vnd.ms-powerpoint',  # .ppt
        ]
        is_office = provided_ct in office_types
        
        # Check if it's already compressed (archives, videos, etc.)
        compressed_types = [
            'application/zip',
            'application/x-zip-compressed',
            'application/x-rar-compressed',
            'application/x-7z-compressed',
            'application/x-tar',
            'application/gzip',
            'video/',
            'audio/',
        ]
        is_already_compressed = any(provided_ct and provided_ct.startswith(ct) for ct in compressed_types)
        
        compressed_data = None
        
        # Try PDF compression if it's a PDF
        if is_pdf:
            try:
                compressed_pdf = compress_pdf(raw)
                if compressed_pdf and len(compressed_pdf) < size:
                    compressed_data = compressed_pdf
                    print(f"PDF compressed: {size} → {len(compressed_data)} bytes ({100 - int(len(compressed_data)/size*100)}% reduction)")
            except Exception as e:
                print(f"PDF compression failed: {e}")
        
        # If not compressed yet and not already a compressed format, try gzip
        if not compressed_data and not is_already_compressed:
            try:
                gzipped = gzip.compress(raw, compresslevel=9)
                # Only use gzip if it actually reduces size significantly (at least 10%)
                if len(gzipped) < size * 0.9:
                    compressed_data = gzipped
                    original_ct = provided_ct or 'application/octet-stream'
                    # Store original content type in filename metadata
                    content_type = 'application/gzip'
                    final_filename = filename + '.gz'
                    print(f"File gzipped: {size} → {len(compressed_data)} bytes ({100 - int(len(compressed_data)/size*100)}% reduction)")
                    
                    # Check if compressed version fits
                    if len(compressed_data) <= MAX_ATTACHMENT_SIZE:
                        return compressed_data, content_type, final_filename, len(compressed_data)
            except Exception as e:
                print(f"Gzip compression failed: {e}")
        
        # Use compressed version if available and smaller
        if compressed_data:
            final_size = len(compressed_data)
            if final_size <= MAX_ATTACHMENT_SIZE:
                return compressed_data, provided_ct or 'application/octet-stream', filename, final_size
        
        # Otherwise use original if it fits
        if size <= MAX_ATTACHMENT_SIZE:
            content_type = provided_ct or 'application/octet-stream'
            return raw, content_type, filename, size

        # File is too large even after compression
        raise ValueError(f'File too large ({size / (1024*1024):.1f}MB). Maximum is 4MB even after compression.')
